"""
ClearAI Dashboard — Phase 2 Backend
FastAPI + SHAP + LIME + XGBoost + GradientBoosting + Groq AI (Llama 3.3 70B)
"""

import uuid
import io
import json
import math
import os
import sqlite3
import warnings
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime, timedelta

from dotenv import load_dotenv
# Load .env.local from the project root (one level up from this backend/ dir)
load_dotenv(Path(__file__).parent.parent / ".env.local")

import joblib

from groq import Groq

import numpy as np
import pandas as pd
from fastapi import FastAPI, File, HTTPException, UploadFile, Query
from fastapi.responses import StreamingResponse

# ── Optional integration packages (graceful fallback if not installed) ────────
try:
    import google.generativeai as genai
    _GEMINI_AVAILABLE = True
except ImportError:
    genai = None  # type: ignore
    _GEMINI_AVAILABLE = False

try:
    import gspread
    from google.oauth2.service_account import Credentials as ServiceAccountCreds
    _GSPREAD_AVAILABLE = True
except ImportError:
    gspread = None  # type: ignore
    ServiceAccountCreds = None  # type: ignore
    _GSPREAD_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None  # type: ignore
    _OPENPYXL_AVAILABLE = False

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.metrics import (
    accuracy_score,
    mean_absolute_error,
    mean_squared_error,
    roc_auc_score,
)
from sklearn.model_selection import cross_val_score, train_test_split
from sklearn.preprocessing import StandardScaler
from xgboost import XGBRegressor, XGBClassifier
import shap
import lime
import lime.lime_tabular

warnings.filterwarnings("ignore")

# ── App ───────────────────────────────────────────────────────────────────
app = FastAPI(title="ClearAI Backend", version="1.0.0")

_DEFAULT_ORIGINS = os.getenv("ALLOWED_ORIGINS", "*")
_origins_raw = _DEFAULT_ORIGINS.strip()
ALLOWED_ORIGINS = ["*"] if _origins_raw == "*" else [o.strip() for o in _origins_raw.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── In-memory caches (warm on first access, populated from DB on restart) ──
DATASETS: Dict[str, Dict] = {}   # dataset_id -> {df, profile, ...}
MODELS: Dict[str, Dict] = {}     # model_id   -> {model, explainer, meta, ...}

# ── SQLite persistence ─────────────────────────────────────────────────────
DB_PATH = Path(__file__).parent / "clearai.db"


def _init_db() -> None:
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS datasets (
            id          TEXT PRIMARY KEY,
            filename    TEXT,
            ext         TEXT,
            uploaded_at TEXT,
            row_count   INTEGER,
            profile     TEXT,
            raw_bytes   BLOB
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS agent_preferences (
            agent_id    TEXT PRIMARY KEY,
            enabled     INTEGER NOT NULL DEFAULT 1,
            updated_at  TEXT
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS agent_activity (
            id          TEXT PRIMARY KEY,
            agent_id    TEXT NOT NULL,
            agent_name  TEXT NOT NULL,
            icon        TEXT NOT NULL,
            action      TEXT NOT NULL,
            output      TEXT,
            status      TEXT NOT NULL DEFAULT 'done',
            created_at  TEXT NOT NULL
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS integration_credentials (
            id           TEXT PRIMARY KEY,
            api_key      TEXT,
            extra        TEXT,
            connected_at TEXT
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS agent_metrics (
            id           TEXT PRIMARY KEY,
            agent_id     TEXT NOT NULL,
            metric_name  TEXT NOT NULL,
            metric_value REAL NOT NULL,
            date         TEXT NOT NULL
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS models (
            id          TEXT PRIMARY KEY,
            goal        TEXT,
            dataset_id  TEXT,
            created_at  TEXT,
            accuracy    REAL,
            metrics     TEXT,
            artifact    BLOB
        )
    """)
    con.commit()
    con.close()


_init_db()


def _save_dataset_to_db(dataset_id: str, entry: Dict) -> None:
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "INSERT OR REPLACE INTO datasets VALUES (?,?,?,?,?,?,?)",
        (
            dataset_id,
            entry["filename"],
            entry["_ext"],
            entry["uploaded_at"],
            len(entry["df"]),
            json.dumps(entry["profile"]),
            entry["_raw"],
        ),
    )
    con.commit()
    con.close()


def _load_dataset_from_db(dataset_id: str) -> Optional[Dict]:
    con = sqlite3.connect(DB_PATH)
    row = con.execute(
        "SELECT filename, ext, uploaded_at, profile, raw_bytes FROM datasets WHERE id=?",
        (dataset_id,),
    ).fetchone()
    con.close()
    if not row:
        return None
    filename, ext, uploaded_at, profile_json, raw_bytes = row
    try:
        df = pd.read_csv(io.BytesIO(raw_bytes)) if ext == "csv" else pd.read_excel(io.BytesIO(raw_bytes))
    except Exception:
        return None
    entry = {
        "df": df,
        "filename": filename,
        "_ext": ext,
        "uploaded_at": uploaded_at,
        "profile": json.loads(profile_json),
        "_raw": raw_bytes,
    }
    DATASETS[dataset_id] = entry
    return entry


def _get_dataset(dataset_id: str) -> Dict:
    if dataset_id in DATASETS:
        return DATASETS[dataset_id]
    entry = _load_dataset_from_db(dataset_id)
    if not entry:
        raise HTTPException(404, "Dataset not found. Please upload your data again.")
    return entry


def _save_model_to_db(model_id: str, entry: Dict) -> None:
    buf = io.BytesIO()
    # Exclude lime_explainer: it contains unpicklable lambda functions internally
    saveable = {k: v for k, v in entry.items() if k != "lime_explainer"}
    joblib.dump(saveable, buf)
    blob = buf.getvalue()
    metrics = {k: entry.get(k) for k in ["accuracy", "auc", "mae", "rmse", "r2"]}
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "INSERT OR REPLACE INTO models VALUES (?,?,?,?,?,?,?)",
        (
            model_id,
            entry["goal"],
            entry.get("dataset_id"),
            datetime.utcnow().isoformat(),
            entry.get("accuracy") or entry.get("r2"),
            json.dumps(metrics),
            blob,
        ),
    )
    con.commit()
    con.close()


def _load_model_from_db(model_id: str) -> Optional[Dict]:
    con = sqlite3.connect(DB_PATH)
    row = con.execute(
        "SELECT artifact FROM models WHERE id=?", (model_id,)
    ).fetchone()
    con.close()
    if not row:
        return None
    try:
        entry = joblib.load(io.BytesIO(row[0]))
        MODELS[model_id] = entry
        return entry
    except Exception:
        return None


def _get_model(model_id: str) -> Dict:
    if model_id in MODELS:
        return MODELS[model_id]
    entry = _load_model_from_db(model_id)
    if not entry:
        raise HTTPException(404, "Model not found. Please retrain your model.")
    return entry

# ── Business-friendly column label mapping ────────────────────────────────
FRIENDLY_LABELS = {
    "days_since_last_purchase": "Days Since Last Purchase",
    "days_inactive": "Days Since Last Purchase",
    "total_spend": "Total Lifetime Spend",
    "lifetime_value": "Total Lifetime Spend",
    "ltv": "Total Lifetime Spend",
    "purchase_count": "Number of Purchases",
    "num_orders": "Number of Purchases",
    "email_open_rate": "Email Engagement Rate",
    "open_rate": "Email Engagement Rate",
    "support_tickets": "Support Ticket Count",
    "complaints": "Support Ticket Count",
    "marketing_spend": "Marketing Spend",
    "is_weekend": "Weekend Flag",
    "day_of_week": "Day of Week",
    "month": "Month of Year",
    "revenue_lag_7": "Revenue 7 Days Ago",
    "revenue_lag_14": "Revenue 14 Days Ago",
    "rolling_mean_7": "7-Day Rolling Avg Revenue",
    "rolling_mean_14": "14-Day Rolling Avg Revenue",
    "revenue": "Revenue",
    "units_sold": "Units Sold",
}


def friendly(col: str) -> str:
    return FRIENDLY_LABELS.get(col, col.replace("_", " ").title())


def safe(v: Any) -> Any:
    """Convert numpy scalar types to Python natives for JSON serialization."""
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return None if math.isnan(v) else float(v)
    if isinstance(v, np.ndarray):
        return [safe(x) for x in v.tolist()]
    if isinstance(v, float) and math.isnan(v):
        return None
    return v


# ── Column auto-detection ─────────────────────────────────────────────────
def detect_column(df: pd.DataFrame, hints: List[str]) -> Optional[str]:
    cols_lower = {c.lower(): c for c in df.columns}
    for hint in hints:
        if hint in cols_lower:
            return cols_lower[hint]
    return None


def detect_date_column(df: pd.DataFrame) -> Optional[str]:
    for col in df.columns:
        if any(k in col.lower() for k in ["date", "time", "day", "period", "week"]):
            try:
                pd.to_datetime(df[col])
                return col
            except Exception:
                pass
    # Try to detect a column that looks like dates
    for col in df.columns:
        try:
            parsed = pd.to_datetime(df[col], infer_datetime_format=True)
            if parsed.notna().sum() > len(df) * 0.8:
                return col
        except Exception:
            pass
    return None


# ── Data profiling ────────────────────────────────────────────────────────
def profile_dataset(df: pd.DataFrame) -> Dict:
    cols = []
    for col in df.columns:
        s = df[col]
        null_pct = round(s.isna().mean() * 100, 1)
        dtype_str = str(s.dtype)
        if "int" in dtype_str or "float" in dtype_str:
            inferred = "numeric"
        elif "datetime" in dtype_str:
            inferred = "date"
        else:
            # Try date detection
            try:
                pd.to_datetime(s.dropna().head(20))
                inferred = "date"
            except Exception:
                inferred = "text" if s.nunique() > 20 else "categorical"

        cols.append({
            "name": col,
            "friendly_name": friendly(col),
            "type": inferred,
            "null_pct": null_pct,
            "unique": int(s.nunique()),
            "sample": [str(v) for v in s.dropna().head(3).tolist()],
        })

    # Quality score
    avg_null = df.isna().mean().mean() * 100
    dup_pct = (df.duplicated().sum() / max(len(df), 1)) * 100
    quality_score = max(0, round(100 - avg_null * 1.5 - dup_pct * 2))

    # Suggest prediction goals
    col_lower = set(c.lower() for c in df.columns)
    suggested = []
    if any(k in col_lower for k in ["revenue", "sales", "amount", "income"]):
        if any(k in col_lower for k in ["date", "time", "day", "period"]):
            suggested.append("sales_forecast")
    if any(k in col_lower for k in ["days_since", "last_purchase", "inactive_days", "days_inactive"]):
        suggested.append("churn_risk")
    if any(k in col_lower for k in ["customer_id", "cust_id"]) and any(
        k in col_lower for k in ["days_since", "purchase_count", "total_spend", "ltv"]
    ):
        if "churn_risk" not in suggested:
            suggested.append("churn_risk")

    return {
        "row_count": len(df),
        "column_count": len(df.columns),
        "quality_score": quality_score,
        "missing_pct": round(avg_null, 1),
        "duplicate_rows": int(df.duplicated().sum()),
        "columns": cols,
        "suggested_goals": suggested if suggested else ["sales_forecast", "churn_risk"],
        "date_range": _date_range(df),
    }


def _date_range(df: pd.DataFrame) -> Optional[str]:
    date_col = detect_date_column(df)
    if not date_col:
        return None
    try:
        dates = pd.to_datetime(df[date_col]).dropna()
        return f"{dates.min().strftime('%b %d, %Y')} — {dates.max().strftime('%b %d, %Y')}"
    except Exception:
        return None


# ── Feature engineering ───────────────────────────────────────────────────
def build_sales_features(df: pd.DataFrame, date_col: str, revenue_col: str) -> pd.DataFrame:
    df = df.copy()
    df[date_col] = pd.to_datetime(df[date_col])
    df = df.sort_values(date_col).reset_index(drop=True)

    df["day_of_week"] = df[date_col].dt.dayofweek
    df["month"] = df[date_col].dt.month
    df["week_of_year"] = df[date_col].dt.isocalendar().week.astype(int)
    df["is_weekend"] = (df["day_of_week"] >= 5).astype(int)
    df["day_of_month"] = df[date_col].dt.day

    # Lag features
    df["revenue_lag_7"] = df[revenue_col].shift(7)
    df["revenue_lag_14"] = df[revenue_col].shift(14)
    df["rolling_mean_7"] = df[revenue_col].shift(1).rolling(7).mean()
    df["rolling_mean_14"] = df[revenue_col].shift(1).rolling(14).mean()

    # Fill NaN lags with median
    for c in ["revenue_lag_7", "revenue_lag_14", "rolling_mean_7", "rolling_mean_14"]:
        df[c] = df[c].fillna(df[revenue_col].median())

    return df


def build_churn_features(df: pd.DataFrame, col_map: Dict) -> pd.DataFrame:
    df = df.copy()
    feature_cols = []

    for feat, hints in {
        "days_since_last_purchase": ["days_since_last_purchase", "days_inactive", "days_since"],
        "total_spend": ["total_spend", "lifetime_value", "ltv", "total_revenue", "amount_spent"],
        "purchase_count": ["purchase_count", "num_orders", "order_count", "num_purchases"],
        "email_open_rate": ["email_open_rate", "open_rate", "email_engagement"],
        "support_tickets": ["support_tickets", "complaints", "ticket_count"],
    }.items():
        if feat in col_map:
            df[feat] = pd.to_numeric(df[col_map[feat]], errors="coerce").fillna(0)
            feature_cols.append(feat)
        else:
            detected = detect_column(df, hints)
            if detected:
                df[feat] = pd.to_numeric(df[detected], errors="coerce").fillna(0)
                feature_cols.append(feat)

    return df, feature_cols


# ── Pydantic models ───────────────────────────────────────────────────────
class TrainRequest(BaseModel):
    dataset_id: str
    goal: str  # "sales_forecast" or "churn_risk"
    column_mapping: Dict[str, str] = {}  # logical_name -> actual_column_name


class SHAPLocalRequest(BaseModel):
    model_id: str
    row_index: int = 0  # index into the scored dataset


class LIMELocalRequest(BaseModel):
    model_id: str
    row_index: int = 0


class CounterfactualRequest(BaseModel):
    model_id: str
    row_index: int = 0
    target_outcome: Optional[float] = None  # desired prediction value/probability


class ExplainSummaryRequest(BaseModel):
    model_id: str
    row_index: int = 0


class NLQueryRequest(BaseModel):
    query: str
    model_id: Optional[str] = None


class ReportRequest(BaseModel):
    model_id: str
    audience: str = "owner"   # "owner" | "investor" | "employee"
    date_range: str = "Last 7 days"


# ── Groq client (lazy — only errors when actually called if key missing) ───
_GROQ_MODEL = "llama-3.3-70b-versatile"


def _groq() -> Groq:
    key = os.getenv("GROQ_API_KEY")
    if not key:
        raise HTTPException(
            503,
            "GROQ_API_KEY is not set. Get a free key at console.groq.com and add it to your environment, then restart the backend."
        )
    return Groq(api_key=key)


def _groq_text(prompt: str, system: str, max_tokens: int = 600) -> str:
    """Call Groq and return the assistant message text."""
    client = _groq()
    completion = client.chat.completions.create(
        model=_GROQ_MODEL,
        max_tokens=max_tokens,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
    )
    return completion.choices[0].message.content.strip()


def _groq_json(prompt: str, system: str, max_tokens: int = 1000) -> str:
    """Call Groq in JSON mode — guaranteed valid JSON output."""
    client = _groq()
    completion = client.chat.completions.create(
        model=_GROQ_MODEL,
        max_tokens=max_tokens,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
    )
    return completion.choices[0].message.content.strip()


# ── Endpoints ─────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "version": "1.0.0"}


@app.post("/api/data/upload")
async def upload_data(file: UploadFile = File(...)):
    """Upload CSV/Excel file. Returns dataset profile and suggested models."""
    if not file.filename:
        raise HTTPException(400, "No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(400, "Only CSV and Excel files are supported")

    contents = await file.read()
    try:
        if ext == "csv":
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    if len(df) < 10:
        raise HTTPException(400, "Dataset must have at least 10 rows")

    dataset_id = str(uuid.uuid4())
    profile = profile_dataset(df)

    entry = {
        "df": df,
        "filename": file.filename,
        "_ext": ext,
        "_raw": contents,
        "uploaded_at": datetime.utcnow().isoformat(),
        "profile": profile,
    }
    DATASETS[dataset_id] = entry
    _save_dataset_to_db(dataset_id, entry)

    return {
        "dataset_id": dataset_id,
        "filename": file.filename,
        "profile": profile,
    }


@app.post("/api/model/train")
def train_model(req: TrainRequest):
    """Train a prediction model for the given goal and dataset."""
    ds = _get_dataset(req.dataset_id)
    df = ds["df"].copy()
    col_map = req.column_mapping

    model_id = str(uuid.uuid4())

    if req.goal == "sales_forecast":
        return _train_sales_forecast(model_id, df, col_map, req.dataset_id)
    elif req.goal == "churn_risk":
        return _train_churn_risk(model_id, df, col_map, req.dataset_id)
    else:
        raise HTTPException(400, f"Unknown goal: {req.goal}. Use 'sales_forecast' or 'churn_risk'")


def _train_sales_forecast(model_id: str, df: pd.DataFrame, col_map: Dict, dataset_id: str) -> Dict:
    # Detect columns
    date_col = col_map.get("date") or detect_date_column(df)
    rev_col = col_map.get("revenue") or detect_column(
        df, ["revenue", "sales", "amount", "total_revenue", "income", "gross_revenue"]
    )

    if not date_col:
        raise HTTPException(400, "Could not detect a date column. Add 'date' to column_mapping.")
    if not rev_col:
        raise HTTPException(400, "Could not detect a revenue column. Add 'revenue' to column_mapping.")

    df_feat = build_sales_features(df, date_col, rev_col)

    feature_cols = [
        "day_of_week", "month", "week_of_year", "is_weekend", "day_of_month",
        "revenue_lag_7", "revenue_lag_14", "rolling_mean_7", "rolling_mean_14",
    ]
    # Include optional numeric columns
    for opt in ["units_sold", "marketing_spend"]:
        mapped = col_map.get(opt) or detect_column(df, [opt])
        if mapped and mapped in df.columns:
            df_feat[opt] = pd.to_numeric(df[mapped], errors="coerce").fillna(0)
            feature_cols.append(opt)

    df_clean = df_feat.dropna(subset=feature_cols + [rev_col])
    X = df_clean[feature_cols].values
    y = pd.to_numeric(df_clean[rev_col], errors="coerce").fillna(0).values

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    model = XGBRegressor(n_estimators=200, learning_rate=0.05, max_depth=4, random_state=42, verbosity=0)
    model.fit(X_train, y_train)

    y_pred = model.predict(X_test)
    mae = float(mean_absolute_error(y_test, y_pred))
    rmse = float(np.sqrt(mean_squared_error(y_test, y_pred)))
    r2 = float(model.score(X_test, y_test))

    # Global SHAP
    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(X_train[:200])  # sample for speed
    global_importance = np.abs(shap_values).mean(axis=0)
    signed_mean = shap_values.mean(axis=0)
    total = global_importance.sum() or 1
    shap_global = [
        {
            "feature": feature_cols[i],
            "label": friendly(feature_cols[i]),
            "importance": round(float(global_importance[i] / total) * 100, 1),
            "raw": float(signed_mean[i]),
        }
        for i in np.argsort(global_importance)[::-1]
    ]

    # Generate future predictions (next 14 days)
    last_date = pd.to_datetime(df_clean[date_col]).max()
    future_rows = []
    last_revenues = y[-14:].tolist() if len(y) >= 14 else y.tolist()

    for i in range(1, 15):
        fd = last_date + timedelta(days=i)
        lag7 = last_revenues[-7] if len(last_revenues) >= 7 else np.median(y)
        lag14 = last_revenues[-14] if len(last_revenues) >= 14 else np.median(y)
        roll7 = float(np.mean(last_revenues[-7:])) if len(last_revenues) >= 7 else float(np.mean(last_revenues))
        roll14 = float(np.mean(last_revenues[-14:])) if len(last_revenues) >= 14 else float(np.mean(last_revenues))
        row = [fd.dayofweek, fd.month, fd.isocalendar()[1], int(fd.dayofweek >= 5), fd.day,
               lag7, lag14, roll7, roll14]
        if "units_sold" in feature_cols:
            row.append(float(np.median(df_feat["units_sold"].fillna(0))))
        if "marketing_spend" in feature_cols:
            row.append(float(np.median(df_feat["marketing_spend"].fillna(0))))
        pred = float(model.predict(np.array([row]))[0])
        future_rows.append({"date": fd.strftime("%b %d"), "predicted": round(pred, 0), "upper": round(pred * 1.12, 0), "lower": round(pred * 0.88, 0)})
        last_revenues.append(pred)

    # Historical actuals
    hist_rows = []
    for _, row in df_clean.tail(14).iterrows():
        hist_rows.append({"date": pd.to_datetime(row[date_col]).strftime("%b %d"), "actual": round(float(row[rev_col]), 0)})

    forecast_series = hist_rows + future_rows

    MODELS[model_id] = {
        "goal": "sales_forecast",
        "model": model,
        "explainer": explainer,
        "feature_cols": feature_cols,
        "shap_global": shap_global,
        "X_train": X_train,
        "X_all": X,
        "y_all": y,
        "df_feat": df_feat,
        "date_col": date_col,
        "rev_col": rev_col,
        "dataset_id": dataset_id,
        "mae": mae,
        "rmse": rmse,
        "r2": r2,
        "forecast_series": forecast_series,
    }
    _save_model_to_db(model_id, MODELS[model_id])

    accuracy_pct = round(max(0, min(100, (1 - mae / (np.mean(y) + 1e-9)) * 100)), 1)

    return {
        "model_id": model_id,
        "goal": "sales_forecast",
        "algorithm": "XGBoost Regressor",
        "accuracy": accuracy_pct,
        "mae": round(mae, 2),
        "rmse": round(rmse, 2),
        "r2": round(r2, 3),
        "training_rows": len(X_train),
        "feature_count": len(feature_cols),
        "shap_global": shap_global,
        "forecast_series": forecast_series,
        "report_card": f"Model trained on {len(X_train)} sales records. Mean absolute error: ${mae:,.0f}/day. The most influential factor is {shap_global[0]['label']} ({shap_global[0]['importance']}% of prediction variance).",
    }


def _train_churn_risk(model_id: str, df: pd.DataFrame, col_map: Dict, dataset_id: str) -> Dict:
    df_feat, feature_cols = build_churn_features(df, col_map)

    if not feature_cols:
        raise HTTPException(400, "Could not detect churn features. Need at least: days_since_last_purchase.")

    # Detect or create target label
    target_col = col_map.get("churned") or detect_column(
        df, ["churned", "churn", "is_churned", "churn_label", "target", "label"]
    )

    if target_col and target_col in df.columns:
        df_feat["_churn_label"] = pd.to_numeric(df[target_col], errors="coerce").fillna(0).astype(int)
    else:
        # Synthesize label: churned if inactive 45+ days
        if "days_since_last_purchase" in feature_cols:
            df_feat["_churn_label"] = (df_feat["days_since_last_purchase"] >= 45).astype(int)
        else:
            raise HTTPException(400, "Need a 'churned' column or 'days_since_last_purchase' to synthesize labels.")

    df_clean = df_feat.dropna(subset=feature_cols).copy()
    X = df_clean[feature_cols].values.astype(float)
    y = df_clean["_churn_label"].values

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y if y.sum() > 5 else None)

    model = XGBClassifier(n_estimators=150, learning_rate=0.05, max_depth=4, random_state=42, verbosity=0, eval_metric="logloss")
    model.fit(X_train, y_train)

    y_pred = model.predict(X_test)
    y_prob = model.predict_proba(X_test)[:, 1]
    accuracy = round(float(accuracy_score(y_test, y_pred)) * 100, 1)
    try:
        auc = round(float(roc_auc_score(y_test, y_prob)), 3)
    except Exception:
        auc = None

    # Global SHAP
    explainer = shap.TreeExplainer(model)
    shap_vals = explainer.shap_values(X_train[:200])
    global_importance = np.abs(shap_vals).mean(axis=0)
    signed_mean = shap_vals.mean(axis=0)
    total = global_importance.sum() or 1
    shap_global = [
        {
            "feature": feature_cols[i],
            "label": friendly(feature_cols[i]),
            "importance": round(float(global_importance[i] / total) * 100, 1),
            "raw": float(signed_mean[i]),
        }
        for i in np.argsort(global_importance)[::-1]
    ]

    # Score all customers
    all_probs = model.predict_proba(X)[:, 1]
    all_shap = explainer.shap_values(X)

    # Build customer records (detect customer_id and name columns)
    id_col = col_map.get("customer_id") or detect_column(df, ["customer_id", "cust_id", "id", "customer"])
    name_col = col_map.get("name") or detect_column(df, ["name", "customer_name", "full_name", "first_name"])

    customers = []
    for idx in np.argsort(all_probs)[::-1][:20]:  # top 20 at-risk
        row = df_clean.iloc[idx]
        cid = str(row[id_col]) if id_col and id_col in row.index else f"C-{idx:04d}"
        cname = str(row[name_col]) if name_col and name_col in row.index else f"Customer {idx + 1}"
        risk = round(float(all_probs[idx]) * 100, 0)

        # Local SHAP for this customer
        sv = all_shap[idx]
        local_shap = sorted(
            [{"feature": feature_cols[i], "label": friendly(feature_cols[i]), "shap_value": safe(sv[i]), "feature_value": safe(float(X[idx][i]))} for i in range(len(feature_cols))],
            key=lambda x: abs(x["shap_value"]), reverse=True,
        )[:5]

        # Revenue at risk
        rev_col_churn = col_map.get("revenue") or detect_column(df, ["total_spend", "revenue", "lifetime_value", "ltv"])
        rev = float(df.iloc[idx][rev_col_churn]) if rev_col_churn and rev_col_churn in df.columns else 0.0

        # Days inactive
        days = int(df_feat.iloc[idx].get("days_since_last_purchase", 0))

        customers.append({
            "id": cid,
            "name": cname,
            "risk": int(risk),
            "revenue": round(rev, 0),
            "days_inactive": days,
            "local_shap": local_shap,
        })

    # LIME explainer (for on-demand use)
    lime_explainer = lime.lime_tabular.LimeTabularExplainer(
        X_train,
        feature_names=[friendly(f) for f in feature_cols],
        class_names=["Active", "Churned"],
        mode="classification",
        discretize_continuous=True,
    )

    MODELS[model_id] = {
        "goal": "churn_risk",
        "model": model,
        "explainer": explainer,
        "lime_explainer": lime_explainer,
        "feature_cols": feature_cols,
        "shap_global": shap_global,
        "X_train": X_train,
        "X_all": X,
        "y_all": y,
        "all_shap": all_shap,
        "customers": customers,
        "dataset_id": dataset_id,
        "accuracy": accuracy,
        "auc": auc,
        "df_feat": df_clean,
    }
    _save_model_to_db(model_id, MODELS[model_id])

    return {
        "model_id": model_id,
        "goal": "churn_risk",
        "algorithm": "XGBoost Classifier",
        "accuracy": accuracy,
        "auc": auc,
        "training_rows": len(X_train),
        "feature_count": len(feature_cols),
        "customers_scored": len(customers),
        "high_risk_count": sum(1 for c in customers if c["risk"] >= 70),
        "shap_global": shap_global,
        "customers": customers,
        "report_card": f"Model trained on {len(X_train)} customer records with {accuracy}% accuracy. {sum(1 for c in customers if c['risk'] >= 70)} customers are at high churn risk (>70%). Top factor: {shap_global[0]['label']}.",
    }


@app.get("/api/model/{model_id}")
def get_model(model_id: str):
    m = _get_model(model_id)
    return {
        "model_id": model_id,
        "goal": m["goal"],
        "accuracy": m.get("accuracy") or m.get("r2"),
        "shap_global": m["shap_global"],
        "forecast_series": m.get("forecast_series"),
        "customers": m.get("customers"),
    }


@app.post("/api/xai/shap/global")
def shap_global(model_id: str):
    return {"shap_global": _get_model(model_id)["shap_global"]}


@app.post("/api/xai/shap/local")
def shap_local(req: SHAPLocalRequest):
    """SHAP waterfall for a single prediction."""
    m = _get_model(req.model_id)
    X = m["X_all"]
    if req.row_index >= len(X):
        raise HTTPException(400, f"Row index {req.row_index} out of bounds ({len(X)} rows)")

    explainer = m["explainer"]
    x = X[req.row_index : req.row_index + 1]
    sv = explainer.shap_values(x)[0]

    try:
        base_value = float(explainer.expected_value)
    except Exception:
        base_value = float(explainer.expected_value[1]) if hasattr(explainer.expected_value, "__len__") else float(explainer.expected_value)

    prediction = float(m["model"].predict(x)[0])
    if m["goal"] == "churn_risk":
        prediction = float(m["model"].predict_proba(x)[0][1])

    waterfall = sorted(
        [{"feature": m["feature_cols"][i], "label": friendly(m["feature_cols"][i]),
          "shap_value": safe(float(sv[i])), "feature_value": safe(float(X[req.row_index][i]))}
         for i in range(len(m["feature_cols"]))],
        key=lambda x: abs(x["shap_value"]), reverse=True,
    )

    return {
        "row_index": req.row_index,
        "base_value": base_value,
        "prediction": prediction,
        "shap_waterfall": waterfall,
    }


@app.post("/api/xai/lime/local")
def lime_local(req: LIMELocalRequest):
    """LIME explanation for a single prediction."""
    m = _get_model(req.model_id)

    if "lime_explainer" not in m:
        # Recreate lazily if it was stripped during DB serialization
        if "X_train" in m and "feature_cols" in m and m.get("goal") == "churn_risk":
            m["lime_explainer"] = lime.lime_tabular.LimeTabularExplainer(
                m["X_train"],
                feature_names=[friendly(f) for f in m["feature_cols"]],
                class_names=["Active", "Churned"],
                mode="classification",
                discretize_continuous=True,
            )
        else:
            raise HTTPException(400, "LIME not available for this model type")

    X = m["X_all"]
    if req.row_index >= len(X):
        raise HTTPException(400, f"Row index out of bounds")

    lime_exp = m["lime_explainer"].explain_instance(
        X[req.row_index],
        m["model"].predict_proba,
        num_features=min(6, len(m["feature_cols"])),
    )

    lime_weights = [
        {"label": name, "weight": safe(weight)}
        for name, weight in lime_exp.as_list()
    ]

    return {
        "row_index": req.row_index,
        "prediction": float(m["model"].predict_proba(X[req.row_index : req.row_index + 1])[0][1]),
        "lime_weights": lime_weights,
    }


@app.post("/api/xai/counterfactual")
def counterfactual(req: CounterfactualRequest):
    """Minimum changes to flip/improve a churn prediction."""
    m = _get_model(req.model_id)

    if m["goal"] != "churn_risk":
        raise HTTPException(400, "Counterfactuals currently supported for churn_risk only")

    X = m["X_all"]
    if req.row_index >= len(X):
        raise HTTPException(400, "Row index out of bounds")

    x_orig = X[req.row_index].copy()
    orig_risk = float(m["model"].predict_proba(x_orig.reshape(1, -1))[0][1])
    target = req.target_outcome if req.target_outcome is not None else 0.5

    feature_cols = m["feature_cols"]
    suggestions = []

    # For each feature, try realistic interventions
    interventions = {
        "days_since_last_purchase": [("make 1 purchase (resets to 1 day)", 1), ("make 1 purchase this week (14 days)", 14)],
        "purchase_count": [("make 1 more purchase", x_orig[feature_cols.index("purchase_count")] + 1 if "purchase_count" in feature_cols else None)],
        "email_open_rate": [("open 3 emails this month (→ 0.30)", 0.30), ("open 5 emails this month (→ 0.50)", 0.50)],
        "support_tickets": [("resolve open tickets (→ 0)", 0)],
    }

    for feat, actions in interventions.items():
        if feat not in feature_cols:
            continue
        fi = feature_cols.index(feat)
        for desc, new_val in actions:
            if new_val is None:
                continue
            x_cf = x_orig.copy()
            x_cf[fi] = float(new_val)
            new_risk = float(m["model"].predict_proba(x_cf.reshape(1, -1))[0][1])
            delta = round((new_risk - orig_risk) * 100, 1)
            if delta < -3:  # meaningful improvement
                suggestions.append({
                    "feature": feat,
                    "label": friendly(feat),
                    "action": desc,
                    "original_value": safe(x_orig[fi]),
                    "new_value": safe(float(new_val)),
                    "original_risk_pct": round(orig_risk * 100, 1),
                    "new_risk_pct": round(new_risk * 100, 1),
                    "delta_pct": delta,
                })

    suggestions.sort(key=lambda x: x["delta_pct"])

    return {
        "row_index": req.row_index,
        "original_risk": round(orig_risk * 100, 1),
        "target_risk": round(target * 100, 1),
        "suggestions": suggestions[:5],
    }


def compute_psi(reference: np.ndarray, current: np.ndarray, bins: int = 10) -> float:
    """Population Stability Index — measures distribution shift between two samples."""
    eps = 1e-6
    if len(reference) < 2 or len(current) < 2:
        return 0.0
    try:
        _, bin_edges = np.histogram(reference, bins=bins)
        bin_edges[0] = -np.inf
        bin_edges[-1] = np.inf
        ref_counts, _ = np.histogram(reference, bins=bin_edges)
        curr_counts, _ = np.histogram(current, bins=bin_edges)
        ref_pct = np.clip(ref_counts / (len(reference) + eps), eps, None)
        curr_pct = np.clip(curr_counts / (len(current) + eps), eps, None)
        psi = float(np.sum((curr_pct - ref_pct) * np.log(curr_pct / ref_pct)))
        return round(max(0.0, psi), 4)
    except Exception:
        return 0.0


@app.get("/api/model/{model_id}/health")
def get_model_health(model_id: str):
    """Return accuracy, per-feature PSI drift, and metadata for a trained model."""
    m = _get_model(model_id)
    X = m["X_all"]
    feature_cols = m["feature_cols"]

    # Split into reference (first 60%) vs current (last 40%) for PSI
    split = max(1, int(len(X) * 0.6))
    ref = X[:split]
    curr = X[split:]

    drift_features = []
    for i, feat in enumerate(feature_cols):
        psi = compute_psi(ref[:, i], curr[:, i])
        status = "none" if psi < 0.1 else "mild" if psi < 0.2 else "moderate"
        drift_features.append({
            "feature": feat,
            "label": friendly(feat),
            "psi": psi,
            "status": status,
        })

    accuracy = m.get("accuracy") or round(float(m.get("r2", 0)) * 100, 1)
    algorithm = "XGBoost Regressor" if m["goal"] == "sales_forecast" else "XGBoost Classifier"

    return {
        "model_id": model_id,
        "goal": m["goal"],
        "algorithm": algorithm,
        "accuracy": accuracy,
        "auc": m.get("auc"),
        "mae": m.get("mae"),
        "r2": m.get("r2"),
        "drift_features": drift_features,
        "feature_count": len(feature_cols),
        "training_rows": len(m["X_train"]),
    }


@app.get("/api/model/{model_id}/fairness")
def get_model_fairness(model_id: str):
    """Return fairness metrics (disparate impact) across customer segments."""
    m = _get_model(model_id)

    if m["goal"] != "churn_risk":
        return {
            "applicable": False,
            "message": "Fairness monitoring applies to classification models (churn risk).",
            "overall_score": 100,
            "attributes": [],
        }

    X = m["X_all"]
    feature_cols = m["feature_cols"]
    y_pred = (m["model"].predict_proba(X)[:, 1] >= 0.5).astype(int)

    attributes = []

    def disparate_impact(groups: list) -> int:
        rates = [g["rate"] for g in groups if g["count"] > 0]
        if not rates or max(rates) == 0:
            return 100
        return min(100, round((min(rates) / max(rates)) * 100))

    # Attribute 1: Activity Level (days_since_last_purchase)
    if "days_since_last_purchase" in feature_cols:
        fi = feature_cols.index("days_since_last_purchase")
        vals = X[:, fi]
        p33, p66 = np.percentile(vals, 33), np.percentile(vals, 66)
        groups = []
        for name, mask in [
            ("Active (recent)", vals <= p33),
            ("Moderate", (vals > p33) & (vals <= p66)),
            ("Inactive (long gap)", vals > p66),
        ]:
            n = int(mask.sum())
            if n > 0:
                rate = round(float(y_pred[mask].mean()), 2)
                groups.append({"name": name, "rate": rate, "count": n})
        if groups:
            score = disparate_impact(groups)
            attributes.append({
                "attr": "Customer Activity Level",
                "score": score,
                "status": "good" if score >= 80 else "warn",
                "groups": groups,
                "note": f"Disparate impact ratio: {score}%. {'Within acceptable range.' if score >= 80 else 'Gap between segments may reflect data imbalance — inactive customers are over-represented in training data.'}",
                "recommendation": None if score >= 80 else "Consider resampling inactive customers or adding recency-adjustment to reduce score gap between segments.",
            })

    # Attribute 2: Spending Tier (total_spend)
    if "total_spend" in feature_cols:
        fi = feature_cols.index("total_spend")
        vals = X[:, fi]
        p50 = np.percentile(vals, 50)
        groups = []
        for name, mask in [("Low Spend", vals <= p50), ("High Spend", vals > p50)]:
            n = int(mask.sum())
            if n > 0:
                rate = round(float(y_pred[mask].mean()), 2)
                groups.append({"name": name, "rate": rate, "count": n})
        if groups:
            score = disparate_impact(groups)
            attributes.append({
                "attr": "Customer Spending Tier",
                "score": score,
                "status": "good" if score >= 80 else "warn",
                "groups": groups,
                "note": f"Churn rate disparity between spending tiers: {score}% disparate impact.",
                "recommendation": None if score >= 80 else "Lower-spend customers appear to churn at higher rates. Consider targeted retention for low-spend segment.",
            })

    # Attribute 3: Email Engagement (email_open_rate)
    if "email_open_rate" in feature_cols:
        fi = feature_cols.index("email_open_rate")
        vals = X[:, fi]
        p50 = np.percentile(vals, 50)
        groups = []
        for name, mask in [("Low Engagement", vals <= p50), ("High Engagement", vals > p50)]:
            n = int(mask.sum())
            if n > 0:
                rate = round(float(y_pred[mask].mean()), 2)
                groups.append({"name": name, "rate": rate, "count": n})
        if groups:
            score = disparate_impact(groups)
            attributes.append({
                "attr": "Email Engagement Level",
                "score": score,
                "status": "good" if score >= 80 else "warn",
                "groups": groups,
                "note": f"Disparate impact ratio: {score}% between engagement segments.",
                "recommendation": None,
            })

    # Fallback: purchase count segments
    if not attributes and "purchase_count" in feature_cols:
        fi = feature_cols.index("purchase_count")
        vals = X[:, fi]
        p50 = np.percentile(vals, 50)
        groups = []
        for name, mask in [("Fewer Purchases", vals <= p50), ("More Purchases", vals > p50)]:
            n = int(mask.sum())
            if n > 0:
                rate = round(float(y_pred[mask].mean()), 2)
                groups.append({"name": name, "rate": rate, "count": n})
        if groups:
            score = disparate_impact(groups)
            attributes.append({
                "attr": "Purchase Frequency",
                "score": score,
                "status": "good" if score >= 80 else "warn",
                "groups": groups,
                "note": f"Disparate impact ratio: {score}% between purchase-frequency segments.",
                "recommendation": None,
            })

    overall_score = round(sum(a["score"] for a in attributes) / len(attributes)) if attributes else 95

    # Build history (simulate 6-week trend based on actual score ± small noise)
    rng = np.random.default_rng(42)
    history = []
    for i, week in enumerate(["W1", "W2", "W3", "W4", "W5", "W6"]):
        entry = {"week": week}
        for j, a in enumerate(attributes):
            noise = int(rng.integers(-4, 5))
            entry[f"score_{j}"] = max(40, min(100, a["score"] - (5 - i) + noise))
        history.append(entry)

    return {
        "applicable": True,
        "overall_score": overall_score,
        "attributes": attributes,
        "history": history,
        "model_id": model_id,
        "goal": m["goal"],
    }


@app.get("/api/model/{model_id}/predictions")
def get_predictions(model_id: str):
    """Return all predictions for a trained model."""
    m = _get_model(model_id)
    if m["goal"] == "sales_forecast":
        return {"goal": "sales_forecast", "forecast_series": m["forecast_series"], "shap_global": m["shap_global"], "mae": m.get("mae"), "rmse": m.get("rmse"), "r2": m.get("r2")}
    else:
        return {"goal": "churn_risk", "customers": m["customers"], "shap_global": m["shap_global"], "accuracy": m.get("accuracy"), "auc": m.get("auc")}


# ── Phase 2: Claude AI endpoints ───────────────────────────────────────────

@app.post("/api/explain/summary")
def explain_summary(req: ExplainSummaryRequest):
    """Ask Claude to explain xAI results in plain English for a business owner."""
    m = _get_model(req.model_id)

    X = m["X_all"]
    if req.row_index >= len(X):
        raise HTTPException(400, "Row index out of bounds")

    # Build SHAP local for the requested row
    explainer = m["explainer"]
    x = X[req.row_index: req.row_index + 1]
    sv = explainer.shap_values(x)[0]
    local_shap = sorted(
        [{"label": friendly(m["feature_cols"][i]), "shap_value": float(sv[i]), "value": float(X[req.row_index][i])}
         for i in range(len(m["feature_cols"]))],
        key=lambda d: abs(d["shap_value"]), reverse=True,
    )[:5]

    # Top global features
    top_global = m["shap_global"][:5]

    # Build context string
    goal = m["goal"]
    accuracy = m.get("accuracy") or round(float(m.get("r2", 0)) * 100, 1)

    if goal == "churn_risk":
        pred_prob = float(m["model"].predict_proba(x)[0][1])
        customers = m.get("customers", [])
        customer = customers[req.row_index] if req.row_index < len(customers) else None
        cname = customer["name"] if customer else f"Customer #{req.row_index}"
        risk_pct = round(pred_prob * 100, 1)

        local_lines = "\n".join(
            f"  - {d['label']}: {'pushes risk UP' if d['shap_value'] > 0 else 'reduces risk'} by {abs(d['shap_value'] * 100):.1f}% (value: {d['value']:.1f})"
            for d in local_shap
        )
        global_lines = "\n".join(
            f"  - {f['label']}: {f['importance']}% of total prediction power"
            for f in top_global
        )

        prompt = f"""A small business owner needs to understand why their AI flagged a customer as at-risk of churning.

Customer: {cname}
Churn risk score: {risk_pct}%
Model accuracy: {accuracy}%

Why this specific customer is at risk (SHAP analysis):
{local_lines}

What drives churn risk across ALL customers globally:
{global_lines}

Write a plain English explanation in exactly 3 short paragraphs:
1. What the AI is saying about {cname} and their churn risk (use the actual percentage)
2. The 2-3 business reasons why they're at risk (translate the technical factors into business language — e.g., "days since last purchase" → "they haven't bought anything in X days")
3. One specific, actionable thing the business owner should do this week to reduce this customer's churn risk

Rules: No jargon (no "SHAP", no "feature importance", no "model"). Write like you're talking to a smart business owner who is good at their trade but not technical. Use plain English. Be direct. Cite specific numbers."""

    else:  # sales_forecast
        pred_val = float(m["model"].predict(x)[0])
        local_lines = "\n".join(
            f"  - {d['label']}: {'increases' if d['shap_value'] > 0 else 'decreases'} forecast by ${abs(d['shap_value']):.0f} (value: {d['value']:.1f})"
            for d in local_shap
        )
        global_lines = "\n".join(
            f"  - {f['label']}: {f['importance']}% of total forecast influence"
            for f in top_global
        )

        prompt = f"""A small business owner needs to understand what their AI sales forecast is telling them.

Predicted revenue for this period: ${pred_val:,.0f}
Model accuracy: {accuracy}%

What's driving this specific forecast (SHAP analysis):
{local_lines}

What drives revenue across all forecasted days globally:
{global_lines}

Write a plain English explanation in exactly 3 short paragraphs:
1. What the AI is predicting and how confident it is
2. The 2-3 biggest business factors driving this forecast (translate technical factors — e.g., "rolling_mean_7" → "your sales momentum from the past week")
3. One specific action the business owner can take to improve on this forecast

Rules: No jargon (no "SHAP", no "features", no "model"). Write like talking to a smart business owner. Be direct. Use specific numbers."""

    system = "You are a friendly business intelligence analyst who explains AI results to small business owners in plain English. You never use technical jargon. You are direct, specific, and always end with an actionable recommendation."

    try:
        text = _groq_text(prompt, system, max_tokens=500)
        return {"summary": text, "goal": goal}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Groq API error: {e}")


@app.post("/api/nl-query")
def nl_query(req: NLQueryRequest):
    """Answer a natural language question about the business using model context."""
    context_lines: List[str] = []

    if req.model_id:
        try:
            m = _get_model(req.model_id)
        except HTTPException:
            m = None
    if req.model_id and m:
        goal = m["goal"]
        accuracy = m.get("accuracy") or round(float(m.get("r2", 0)) * 100, 1)
        context_lines.append(f"Active AI model: {goal.replace('_', ' ').title()} (accuracy: {accuracy}%)")
        context_lines.append(f"Algorithm: {'XGBoost Regressor' if goal == 'sales_forecast' else 'XGBoost Classifier'}")

        top_features = ", ".join(f["label"] for f in m["shap_global"][:3])
        context_lines.append(f"Top prediction drivers: {top_features}")

        if goal == "sales_forecast":
            fs = m.get("forecast_series", [])
            actuals = [p["actual"] for p in fs if p.get("actual") is not None]
            preds = [p["predicted"] for p in fs if p.get("predicted") is not None]
            if actuals:
                context_lines.append(f"Recent revenue (last {len(actuals)} days): ${sum(actuals):,.0f} total, avg ${sum(actuals)/len(actuals):,.0f}/day")
            if preds:
                context_lines.append(f"AI forecasted revenue (next {len(preds)} days): ${sum(preds):,.0f}")
            if m.get("mae"):
                context_lines.append(f"Forecast error margin: ±${m['mae']:,.0f}/day (MAE)")

        elif goal == "churn_risk":
            customers = m.get("customers", [])
            high_risk = sum(1 for c in customers if c["risk"] >= 70)
            top_names = ", ".join(c["name"] for c in customers[:3])
            context_lines.append(f"Customers scored: {len(customers)}")
            context_lines.append(f"High churn risk (>70%): {high_risk} customers")
            context_lines.append(f"Most at-risk: {top_names}")
            if m.get("auc"):
                context_lines.append(f"Model AUC: {m['auc']:.3f}")

    context = "\n".join(context_lines) if context_lines else "No model has been trained yet."

    system = (
        "You are an AI business analyst for a small business. "
        "Answer questions about business performance and AI model results in plain English. "
        "Be direct, specific, and always use numbers from the context. "
        "Do not make up numbers not in the context. "
        "If a question can't be answered from the context, say what you know and what additional data would help. "
        "Keep answers to 2-3 sentences maximum."
    )

    prompt = f"""Business data context:
{context}

Business owner's question: "{req.query}"

Answer in 2-3 sentences. Be specific, use numbers, be actionable."""

    try:
        answer = _groq_text(prompt, system, max_tokens=250)
        return {"answer": answer}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Groq API error: {e}")


@app.post("/api/reports/generate")
def generate_report(req: ReportRequest):
    """Generate a full AI business brief using Claude, grounded in real model data."""
    m = _get_model(req.model_id)

    goal = m["goal"]
    accuracy = m.get("accuracy") or round(float(m.get("r2", 0)) * 100, 1)
    top_features = m["shap_global"][:4]
    top_feature_names = ", ".join(f["label"] for f in top_features)

    data_block: List[str] = [
        f"Model type: {goal.replace('_', ' ').title()}",
        f"Accuracy: {accuracy}%",
        f"Date range covered: {req.date_range}",
        f"Top prediction drivers: {top_feature_names}",
    ]

    if goal == "sales_forecast":
        fs = m.get("forecast_series", [])
        actuals = [(p["date"], p["actual"]) for p in fs if p.get("actual") is not None]
        preds = [(p["date"], p["predicted"]) for p in fs if p.get("predicted") is not None]
        if actuals:
            total_actual = sum(v for _, v in actuals)
            avg_actual = total_actual / len(actuals)
            data_block.append(f"Revenue over last {len(actuals)} days: ${total_actual:,.0f} total (avg ${avg_actual:,.0f}/day)")
        if preds:
            total_pred = sum(v for _, v in preds)
            data_block.append(f"AI forecast next {len(preds)} days: ${total_pred:,.0f}")
        if m.get("mae"):
            data_block.append(f"Forecast error margin: ±${m['mae']:,.0f}/day")
        if m.get("r2"):
            data_block.append(f"R² score: {m['r2']:.3f}")

    elif goal == "churn_risk":
        customers = m.get("customers", [])
        high_risk = [c for c in customers if c["risk"] >= 70]
        med_risk = [c for c in customers if 50 <= c["risk"] < 70]
        at_risk_rev = sum(c.get("revenue", 0) for c in high_risk)
        data_block.append(f"Total customers scored: {len(customers)}")
        data_block.append(f"High churn risk (>70%): {len(high_risk)} customers")
        data_block.append(f"Medium churn risk (50-70%): {len(med_risk)} customers")
        if at_risk_rev > 0:
            data_block.append(f"Revenue at risk from high-risk customers: ${at_risk_rev:,.0f}")
        if customers:
            top3 = customers[:3]
            top3_str = ", ".join(f"{c['name']} ({c['risk']}%)" for c in top3)
            data_block.append(f"Top 3 at-risk customers: {top3_str}")
        if m.get("auc"):
            data_block.append(f"Model AUC: {m['auc']:.3f}")

    audience_instruction = {
        "owner": "Write for the business owner who runs day-to-day operations. Focus on what to do now and revenue impact.",
        "investor": "Write for a business investor or board member. Focus on performance trends, risk quantification, and strategic implications.",
        "employee": "Write for a team member who will execute recommendations. Focus on clear, specific tasks they can take action on.",
    }.get(req.audience, "Write for the business owner.")

    data_str = "\n".join(f"  - {line}" for line in data_block)

    prompt = f"""Generate a weekly AI business intelligence brief. Use ONLY the data below — do not invent numbers.

{audience_instruction}

Business data:
{data_str}

Return a single valid JSON object with this exact schema (no other text, no markdown, just JSON):
{{
  "headline": "One compelling sentence with a key number (max 80 chars)",
  "summary": "2-3 sentence executive summary: main finding, biggest risk, and top opportunity. Use specific numbers.",
  "insights": [
    {{
      "title": "Short insight title with a number",
      "body": "2-3 sentences. Use specific numbers from the data. Explain business impact, not technical details.",
      "driver": "Key driver phrase (5-10 words)",
      "sentiment": "positive"
    }}
  ],
  "recommendations": [
    "Specific, actionable recommendation with timeline. Start with a verb. Include expected outcome."
  ]
}}

Generate exactly 3 insights and 3 recommendations. sentiment must be one of: positive, warning, negative."""

    system = (
        "You are an expert business intelligence analyst generating structured reports for small businesses. "
        "You write clearly, cite specific numbers, and always make recommendations actionable. "
        "You return only valid JSON — no markdown, no preamble, no explanation outside the JSON."
    )

    try:
        raw = _groq_json(prompt, system, max_tokens=1000)
        report = json.loads(raw)
        report["generated_at"] = datetime.utcnow().strftime("%a %b %d, %Y · %I:%M %p UTC")
        report["date_range"] = req.date_range
        report["audience"] = req.audience
        return report
    except json.JSONDecodeError as e:
        raise HTTPException(500, f"Groq returned malformed JSON: {e}. Raw: {raw[:200]}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Groq API error: {e}")


# ── Persistence meta-endpoints ─────────────────────────────────────────────

@app.get("/api/models")
def list_models():
    """List all persisted models with metadata."""
    con = sqlite3.connect(DB_PATH)
    rows = con.execute(
        "SELECT id, goal, dataset_id, created_at, accuracy, metrics FROM models ORDER BY created_at DESC"
    ).fetchall()
    con.close()
    return [
        {
            "model_id": r[0],
            "goal": r[1],
            "dataset_id": r[2],
            "created_at": r[3],
            "accuracy": r[4],
            "metrics": json.loads(r[5]) if r[5] else {},
        }
        for r in rows
    ]


@app.get("/api/datasets")
def list_datasets():
    """List all persisted datasets."""
    con = sqlite3.connect(DB_PATH)
    rows = con.execute(
        "SELECT id, filename, uploaded_at, row_count FROM datasets ORDER BY uploaded_at DESC"
    ).fetchall()
    con.close()
    return [
        {"dataset_id": r[0], "filename": r[1], "uploaded_at": r[2], "row_count": r[3]}
        for r in rows
    ]


@app.delete("/api/model/{model_id}")
def delete_model(model_id: str):
    """Remove a model from both cache and DB."""
    MODELS.pop(model_id, None)
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM models WHERE id=?", (model_id,))
    con.commit()
    con.close()
    return {"deleted": model_id}


# ── Agent infrastructure ───────────────────────────────────────────────────

# Default enabled state per agent (active agents on by default, idle ones off)
_AGENT_DEFAULTS = {
    "marketing": True, "cx": True, "sales": True, "finance": True,
    "hr": False, "ops": True, "bi": True, "inventory": False,
    "reputation": True, "seo": True, "compliance": False, "comms": True,
}

# Per-agent configuration: name, icon, color, system prompt, task prompt template
AGENT_CONFIGS: Dict[str, Dict] = {
    "marketing": {
        "name": "Marketing Agent", "icon": "🎯", "color": "#4F46FF",
        "system": (
            "You are an expert AI marketing agent for a small business. "
            "You generate specific, actionable marketing content including actual post copy, "
            "email subject lines, and campaign strategies. Always produce concrete, ready-to-use output."
        ),
        "prompt_template": (
            "You just ran a marketing cycle for {business_context}. "
            "Generate exactly 3 completed tasks. For each task produce the actual content output "
            "(e.g. the real social post text, the actual email subject line, the specific ad headline). "
            "Return ONLY valid JSON matching this schema exactly:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<actual content produced>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "cx": {
        "name": "CX Agent", "icon": "🎧", "color": "#00E5A0",
        "system": (
            "You are a 24/7 AI customer experience agent. "
            "You handle customer inquiries, resolve issues, and generate response templates. "
            "Be empathetic, specific, and professional."
        ),
        "prompt_template": (
            "You just completed a customer service shift for {business_context}. "
            "Generate 3 tasks showing what you handled: resolved inquiries, drafted responses, "
            "or escalation summaries. Include the actual response text in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<actual response/action>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "sales": {
        "name": "Sales & CRM Agent", "icon": "💼", "color": "#FFC85C",
        "system": (
            "You are an AI sales and CRM agent. You manage pipelines, score leads, "
            "write follow-up sequences, and update CRM records automatically. "
            "Be specific with deal names, values, and next steps."
        ),
        "prompt_template": (
            "You just ran a sales pipeline review for {business_context}. "
            "Generate 3 tasks: e.g. lead scoring results, a follow-up email draft, pipeline summary. "
            "Include actual email copy or specific deal data in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<actual output>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "finance": {
        "name": "Finance & Accounting Agent", "icon": "💰", "color": "#00D4FF",
        "system": (
            "You are an AI finance and accounting agent for a small business. "
            "You analyze transactions, generate P&L summaries, flag anomalies, and forecast cash flow. "
            "Use realistic numbers and plain-language explanations."
        ),
        "prompt_template": (
            "You just ran a financial analysis for {business_context}. "
            "Generate 3 tasks: e.g. transaction categorization, cash flow forecast, anomaly detection. "
            "Include specific numbers and findings in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<findings with numbers>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "hr": {
        "name": "HR & People Ops Agent", "icon": "👥", "color": "#FF6B9D",
        "system": (
            "You are an AI HR and people operations agent. "
            "You screen applicants, draft job postings, manage onboarding checklists, "
            "and handle scheduling. Be practical and specific."
        ),
        "prompt_template": (
            "You just ran an HR operations cycle for {business_context}. "
            "Generate 3 tasks: e.g. job posting draft, applicant screening summary, onboarding checklist. "
            "Include the actual job post text or checklist items in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<actual content>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "ops": {
        "name": "Operations & Workflow Agent", "icon": "⚙️", "color": "#9B8BFF",
        "system": (
            "You are an AI operations and workflow agent. "
            "You identify bottlenecks, automate task assignment, generate SOPs, "
            "and orchestrate multi-step business workflows."
        ),
        "prompt_template": (
            "You just ran an operations audit for {business_context}. "
            "Generate 3 tasks: e.g. bottleneck identified, SOP drafted, task assignments updated. "
            "Include specific workflow steps or SOP content in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<specific output>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "bi": {
        "name": "Business Intelligence Agent", "icon": "📊", "color": "#4F46FF",
        "system": (
            "You are an AI business intelligence analyst. "
            "You consolidate data, identify trends, generate KPI reports, and deliver actionable forecasts. "
            "Always ground insights in specific metrics and numbers."
        ),
        "prompt_template": (
            "You just ran a business intelligence cycle for {business_context}. "
            "Generate 3 tasks: e.g. weekly KPI report, trend analysis, forecast update. "
            "Include specific metrics, percentages, and findings in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<report with numbers>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "inventory": {
        "name": "Inventory & Supply Chain Agent", "icon": "📦", "color": "#FF8C42",
        "system": (
            "You are an AI inventory and supply chain agent. "
            "You track stock levels, generate purchase orders, forecast demand, "
            "and flag supply chain risks. Use specific SKUs and quantities."
        ),
        "prompt_template": (
            "You just ran an inventory check for {business_context}. "
            "Generate 3 tasks: e.g. low-stock alert, purchase order generated, demand forecast. "
            "Include specific products, quantities, and reorder points in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<specific inventory data>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "reputation": {
        "name": "Reputation & Review Agent", "icon": "⭐", "color": "#FFD700",
        "system": (
            "You are an AI reputation management agent. "
            "You monitor reviews across Google, Yelp, TripAdvisor, and social platforms, "
            "draft personalized responses, and track NPS trends."
        ),
        "prompt_template": (
            "You just ran a reputation management cycle for {business_context}. "
            "Generate 3 tasks: e.g. new review found and response drafted, NPS update, sentiment trend. "
            "Include the actual review response text in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<actual response draft>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "seo": {
        "name": "SEO & Content Agent", "icon": "🔍", "color": "#34D399",
        "system": (
            "You are an AI SEO and content strategy agent. "
            "You perform keyword research, audit page health, draft blog post outlines, "
            "and monitor ranking changes. Be specific with keyword data and recommendations."
        ),
        "prompt_template": (
            "You just ran an SEO cycle for {business_context}. "
            "Generate 3 tasks: e.g. keyword opportunities found, blog post outline drafted, SEO audit. "
            "Include specific keywords with estimated search volume and the actual outline in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<keywords or outline>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "compliance": {
        "name": "Compliance & Risk Agent", "icon": "🛡️", "color": "#60A5FA",
        "system": (
            "You are an AI compliance and risk management agent. "
            "You monitor contracts, track regulatory deadlines, flag data privacy risks, "
            "and maintain audit trails. Be precise and risk-aware."
        ),
        "prompt_template": (
            "You just ran a compliance check for {business_context}. "
            "Generate 3 tasks: e.g. contract review, regulatory update flagged, GDPR checklist. "
            "Include specific risks, deadlines, and checklist items in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<specific compliance findings>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
    "comms": {
        "name": "Communication & Email Agent", "icon": "📧", "color": "#F472B6",
        "system": (
            "You are an AI communication and email management agent. "
            "You triage inboxes, draft context-aware emails, summarize threads, "
            "and manage newsletter content. Always produce the actual email text."
        ),
        "prompt_template": (
            "You just ran an inbox management cycle for {business_context}. "
            "Generate 3 tasks: e.g. inbox triaged, email draft written, thread summarized. "
            "Include the actual email draft or summary text in outputs. "
            "Return ONLY valid JSON:\n"
            '{{"summary":"<one sentence>","tasks":[{{"title":"<task name>","description":"<what was done>","output":"<actual email or summary>","status":"done"}}],"insights":["<insight1>","<insight2>"],"action_required":false,"action_description":null}}'
        ),
    },
}


class AgentToggleRequest(BaseModel):
    enabled: bool


class AgentRunRequest(BaseModel):
    context: Optional[Dict[str, Any]] = None  # model data, business info


def _get_agent_preferences() -> Dict[str, bool]:
    con = sqlite3.connect(DB_PATH)
    rows = con.execute("SELECT agent_id, enabled FROM agent_preferences").fetchall()
    con.close()
    prefs = dict(_AGENT_DEFAULTS)
    for agent_id, enabled in rows:
        prefs[agent_id] = bool(enabled)
    return prefs


def _save_activity(agent_id: str, action: str, output: str, status: str = "done") -> Dict:
    cfg = AGENT_CONFIGS.get(agent_id, {})
    entry = {
        "id": str(uuid.uuid4()),
        "agent_id": agent_id,
        "agent_name": cfg.get("name", agent_id),
        "icon": cfg.get("icon", "🤖"),
        "color": cfg.get("color", "#4F46FF"),
        "action": action,
        "output": output,
        "status": status,
        "created_at": datetime.utcnow().isoformat(),
    }
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "INSERT INTO agent_activity VALUES (?,?,?,?,?,?,?,?)",
        (entry["id"], agent_id, entry["agent_name"], entry["icon"],
         action, output, status, entry["created_at"]),
    )
    con.commit()
    con.close()
    return entry


@app.get("/api/agents")
def get_agents():
    """Return all agent configs with current preferences and task counts."""
    prefs = _get_agent_preferences()
    con = sqlite3.connect(DB_PATH)
    today = datetime.utcnow().strftime("%Y-%m-%d")
    counts = {
        row[0]: row[1]
        for row in con.execute(
            "SELECT agent_id, COUNT(*) FROM agent_activity WHERE created_at >= ? GROUP BY agent_id",
            (today,),
        ).fetchall()
    }
    last_runs = {
        row[0]: row[1]
        for row in con.execute(
            "SELECT agent_id, MAX(created_at) FROM agent_activity GROUP BY agent_id"
        ).fetchall()
    }
    con.close()

    agents = []
    for agent_id, cfg in AGENT_CONFIGS.items():
        last_raw = last_runs.get(agent_id)
        if last_raw:
            delta = datetime.utcnow() - datetime.fromisoformat(last_raw)
            if delta.total_seconds() < 60:
                last_run_str = "just now"
            elif delta.total_seconds() < 3600:
                last_run_str = f"{int(delta.total_seconds() / 60)}m ago"
            elif delta.total_seconds() < 86400:
                last_run_str = f"{int(delta.total_seconds() / 3600)}h ago"
            else:
                last_run_str = f"{int(delta.total_seconds() / 86400)}d ago"
        else:
            last_run_str = "never"

        agents.append({
            "id": agent_id,
            "name": cfg["name"],
            "icon": cfg["icon"],
            "color": cfg["color"],
            "enabled": prefs.get(agent_id, _AGENT_DEFAULTS.get(agent_id, True)),
            "tasks_today": counts.get(agent_id, 0),
            "last_run": last_run_str,
        })
    return agents


@app.post("/api/agents/{agent_id}/toggle")
def toggle_agent(agent_id: str, req: AgentToggleRequest):
    if agent_id not in AGENT_CONFIGS:
        raise HTTPException(404, f"Unknown agent: {agent_id}")
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "INSERT OR REPLACE INTO agent_preferences VALUES (?,?,?)",
        (agent_id, 1 if req.enabled else 0, datetime.utcnow().isoformat()),
    )
    con.commit()
    con.close()
    return {"agent_id": agent_id, "enabled": req.enabled}


def _extract_agent_metrics(agent_id: str, result: Dict) -> Dict[str, float]:
    """Parse numeric KPIs out of an agent run result for time-series tracking."""
    import re as _re
    tasks = result.get("tasks", [])
    all_text = " ".join(
        (t.get("output", "") + " " + t.get("description", "") + " " + t.get("title", ""))
        for t in tasks
    ).lower()

    def _first_num(text: str, lo: int = 1, hi: int = 9999, default: float = 0.0) -> float:
        candidates = [int(n) for n in _re.findall(r'\b(\d{1,5})\b', text) if lo <= int(n) <= hi]
        return float(candidates[0]) if candidates else default

    if agent_id == "marketing":
        big_nums = [int(n) for n in _re.findall(r'\b(\d{3,6})\b', all_text) if 100 <= int(n) <= 999999]
        reach = float(big_nums[0]) if big_nums else float(len(tasks)) * 420
        return {"engagement_reach": reach}
    elif agent_id == "cx":
        nums = _re.findall(r'(?:resolved|handled|closed|answered)\D{0,5}(\d+)|(\d+)\D{0,5}(?:ticket|inquir|issue|customer)', all_text)
        flat = [int(n) for pair in nums for n in pair if n]
        return {"tickets_resolved": float(flat[0]) if flat else _first_num(all_text, 1, 50, 8.0)}
    elif agent_id == "sales":
        deal_nums = _re.findall(r'(\d+)\D{0,5}(?:lead|deal|prospect|opport)', all_text)
        return {"leads_scored": float(deal_nums[0]) if deal_nums else _first_num(all_text, 1, 30, 5.0)}
    elif agent_id == "seo":
        kw_nums = _re.findall(r'(\d+)\D{0,5}(?:keyword|phrase|term)', all_text)
        return {"keywords_tracked": float(kw_nums[0]) if kw_nums else _first_num(all_text, 1, 50, 7.0)}
    elif agent_id == "reputation":
        rev_nums = _re.findall(r'(\d+)\D{0,5}(?:review|rating|feedback)', all_text)
        return {"reviews_monitored": float(rev_nums[0]) if rev_nums else _first_num(all_text, 1, 20, 3.0)}
    elif agent_id == "finance":
        return {"transactions_reviewed": _first_num(all_text, 5, 200, 45.0)}
    elif agent_id == "bi":
        return {"kpis_tracked": _first_num(all_text, 1, 30, 8.0)}
    elif agent_id == "ops":
        return {"tasks_automated": _first_num(all_text, 1, 20, 6.0)}
    elif agent_id == "comms":
        email_nums = _re.findall(r'(\d+)\D{0,5}(?:email|message|draft)', all_text)
        return {"emails_drafted": float(email_nums[0]) if email_nums else _first_num(all_text, 1, 20, 5.0)}
    elif agent_id == "compliance":
        risk_nums = _re.findall(r'(\d+)\D{0,5}(?:risk|issue|violation|concern)', all_text)
        return {"risks_identified": float(risk_nums[0]) if risk_nums else 0.0}
    elif agent_id == "hr":
        cand_nums = _re.findall(r'(\d+)\D{0,5}(?:candidate|applicant|resume)', all_text)
        return {"candidates_screened": float(cand_nums[0]) if cand_nums else _first_num(all_text, 1, 15, 4.0)}
    elif agent_id == "inventory":
        item_nums = _re.findall(r'(\d+)\D{0,5}(?:item|sku|product|unit)', all_text)
        return {"items_checked": float(item_nums[0]) if item_nums else _first_num(all_text, 5, 500, 20.0)}
    return {"tasks_completed": float(len(tasks))}


def _save_agent_metrics(agent_id: str, metrics: Dict[str, float]) -> None:
    today = datetime.utcnow().strftime("%Y-%m-%d")
    con = sqlite3.connect(DB_PATH)
    for metric_name, metric_value in metrics.items():
        con.execute(
            "INSERT OR REPLACE INTO agent_metrics VALUES (?,?,?,?,?)",
            (str(uuid.uuid4()), agent_id, metric_name, metric_value, today),
        )
    con.commit()
    con.close()


@app.post("/api/agents/{agent_id}/run")
def run_agent(agent_id: str, req: AgentRunRequest):
    """Run an agent using Groq and persist the activity."""
    if agent_id not in AGENT_CONFIGS:
        raise HTTPException(404, f"Unknown agent: {agent_id}")

    cfg = AGENT_CONFIGS[agent_id]
    ctx = req.context or {}

    # Build business context string from whatever data is available
    business_parts = []
    if ctx.get("goal") == "sales_forecast":
        business_parts.append(f"a retail/sales business with {ctx.get('training_rows', 'unknown')} records")
        if ctx.get("accuracy"):
            business_parts.append(f"current sales forecast accuracy of {ctx['accuracy']}%")
        if ctx.get("actual_rev"):
            business_parts.append(f"recent revenue of ${ctx['actual_rev']:,.0f}")
    elif ctx.get("goal") == "churn_risk":
        business_parts.append(f"a subscription/service business with {ctx.get('customers_scored', 'unknown')} customers scored")
        if ctx.get("high_risk_count"):
            business_parts.append(f"{ctx['high_risk_count']} high-risk customers identified")
    if not business_parts:
        business_parts.append("a small business (no model data available yet)")

    business_context = "; ".join(business_parts)
    prompt = cfg["prompt_template"].format(business_context=business_context)

    try:
        raw = _groq_json(prompt, cfg["system"], max_tokens=900)
        result = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(500, "Agent returned malformed output. Please try again.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Agent error: {e}")

    # ── Nano Banana Pro (Gemini) augmentation for creative agents ──────────────
    GEMINI_AGENTS = {"marketing", "seo", "reputation", "comms"}
    if agent_id in GEMINI_AGENTS and _GEMINI_AVAILABLE:
        gemini_cred = _get_integration("nano_banana_pro")
        if gemini_cred and gemini_cred.get("api_key"):
            try:
                genai.configure(api_key=gemini_cred["api_key"])
                gm = genai.GenerativeModel("gemini-1.5-flash")
                creative_prompts = {
                    "marketing": f"Generate 3 ready-to-use marketing assets for {business_context}: (1) a social post with hashtags, (2) an email subject + preview text, (3) a paid ad headline + description. Be specific, creative, and brand-appropriate.",
                    "seo": f"Generate SEO content for {business_context}: (1) a blog post title + opening paragraph optimized for search, (2) 5 long-tail keywords with rationale, (3) a homepage meta description.",
                    "reputation": f"Generate reputation responses for {business_context}: (1) a warm response to a 5-star review, (2) a professional recovery response to a 2-star complaint, (3) a template for the most common customer issue.",
                    "comms": f"Generate communication drafts for {business_context}: (1) a monthly newsletter intro, (2) a 5-day follow-up email, (3) a team update announcement. Be clear, warm, and professional.",
                }
                gemini_resp = gm.generate_content(creative_prompts.get(agent_id, f"Generate creative business content for: {business_context}"))
                result.setdefault("tasks", []).append({
                    "title": "Creative Assets · Nano Banana Pro",
                    "description": "AI-generated creative content powered by Google Gemini",
                    "output": gemini_resp.text,
                    "status": "done",
                })
                result["gemini_powered"] = True
            except Exception:
                pass  # Don't fail the whole run if Gemini errors

    # Persist to activity log
    summary = result.get("summary", "Task completed")
    tasks = result.get("tasks", [])
    first_task_output = tasks[0]["output"] if tasks else ""
    _save_activity(agent_id, summary, first_task_output, "done")

    result["agent_id"] = agent_id
    result["agent_name"] = cfg["name"]
    result["icon"] = cfg["icon"]
    result["color"] = cfg["color"]
    result["ran_at"] = datetime.utcnow().isoformat()

    # Persist per-agent performance metrics for analytics charts
    try:
        _save_agent_metrics(agent_id, _extract_agent_metrics(agent_id, result))
    except Exception:
        pass  # Never block a run for analytics logging

    return result


@app.get("/api/agents/activity")
def get_agent_activity(limit: int = 20):
    """Return recent agent activity entries."""
    con = sqlite3.connect(DB_PATH)
    rows = con.execute(
        "SELECT id, agent_id, agent_name, icon, action, output, status, created_at "
        "FROM agent_activity ORDER BY created_at DESC LIMIT ?",
        (limit,),
    ).fetchall()
    con.close()

    result = []
    for row in rows:
        _id, agent_id, agent_name, icon, action, output, status, created_at = row
        cfg = AGENT_CONFIGS.get(agent_id, {})
        delta = datetime.utcnow() - datetime.fromisoformat(created_at)
        if delta.total_seconds() < 60:
            time_str = "just now"
        elif delta.total_seconds() < 3600:
            time_str = f"{int(delta.total_seconds() / 60)}m ago"
        elif delta.total_seconds() < 86400:
            time_str = f"{int(delta.total_seconds() / 3600)}h ago"
        else:
            time_str = f"{int(delta.total_seconds() / 86400)}d ago"

        result.append({
            "id": _id,
            "agent_id": agent_id,
            "agent": agent_name,
            "icon": icon,
            "color": cfg.get("color", "#4F46FF"),
            "action": action,
            "output": output,
            "status": status,
            "time": time_str,
        })
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# ── Integration Infrastructure ─────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

# ── DB helpers ────────────────────────────────────────────────────────────────

def _get_integration(integration_id: str) -> Optional[Dict]:
    con = sqlite3.connect(DB_PATH)
    row = con.execute(
        "SELECT api_key, extra, connected_at FROM integration_credentials WHERE id=?",
        (integration_id,),
    ).fetchone()
    con.close()
    if not row:
        return None
    return {"api_key": row[0], "extra": row[1], "connected_at": row[2]}


def _save_integration(integration_id: str, api_key: Optional[str], extra: Optional[str]) -> None:
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "INSERT OR REPLACE INTO integration_credentials VALUES (?,?,?,?)",
        (integration_id, api_key, extra, datetime.utcnow().isoformat()),
    )
    con.commit()
    con.close()


def _delete_integration(integration_id: str) -> None:
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM integration_credentials WHERE id=?", (integration_id,))
    con.commit()
    con.close()


def _get_all_connected() -> List[str]:
    con = sqlite3.connect(DB_PATH)
    rows = con.execute("SELECT id FROM integration_credentials").fetchall()
    con.close()
    return [r[0] for r in rows]


# ── Pydantic models ────────────────────────────────────────────────────────────

class IntegrationConnectRequest(BaseModel):
    api_key: Optional[str] = None
    extra: Optional[Dict[str, Any]] = None


class GeminiGenerateRequest(BaseModel):
    prompt: str
    agent_id: Optional[str] = None


class SheetsExportRequest(BaseModel):
    model_id: str
    spreadsheet_id: Optional[str] = None
    sheet_name: str = "FlowDesk Predictions"


class SlackMessageRequest(BaseModel):
    text: str
    blocks: Optional[List[Dict]] = None


# ── Integration CRUD endpoints ─────────────────────────────────────────────────

@app.get("/api/integrations")
def list_integration_status():
    """Return list of connected integration IDs."""
    return {"connected": _get_all_connected()}


@app.post("/api/integrations/{integration_id}/connect")
def connect_integration(integration_id: str, req: IntegrationConnectRequest):
    """Store credentials for an integration."""
    if not req.api_key and not req.extra:
        raise HTTPException(400, "At least one credential field is required.")
    extra_json = json.dumps(req.extra) if req.extra else None
    _save_integration(integration_id, req.api_key, extra_json)
    return {"integration_id": integration_id, "connected": True}


@app.post("/api/integrations/{integration_id}/disconnect")
def disconnect_integration(integration_id: str):
    """Remove stored credentials."""
    _delete_integration(integration_id)
    return {"integration_id": integration_id, "connected": False}


# ── Nano Banana Pro — Google Gemini ───────────────────────────────────────────

@app.post("/api/integrations/gemini/generate")
def gemini_generate(req: GeminiGenerateRequest):
    """Generate content via Nano Banana Pro (Google Gemini 1.5 Flash)."""
    if not _GEMINI_AVAILABLE:
        raise HTTPException(503, "google-generativeai not installed. Run: pip install google-generativeai")
    cred = _get_integration("nano_banana_pro")
    if not cred or not cred.get("api_key"):
        raise HTTPException(400, "Nano Banana Pro not connected. Add your Gemini API key in the Integrations hub.")
    genai.configure(api_key=cred["api_key"])
    model = genai.GenerativeModel("gemini-1.5-flash")
    try:
        resp = model.generate_content(req.prompt)
        return {"content": resp.text, "model": "gemini-1.5-flash", "powered_by": "Nano Banana Pro"}
    except Exception as e:
        raise HTTPException(500, f"Gemini API error: {e}")


# ── Excel Export ───────────────────────────────────────────────────────────────

@app.get("/api/integrations/excel/export")
def export_excel(model_id: str):
    """Export model predictions to a formatted .xlsx file."""
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(503, "openpyxl not installed. Run: pip install openpyxl")

    m = _get_model(model_id)
    wb = Workbook()

    BRAND_PURPLE = "4F46FF"
    HEADER_FILL = PatternFill(start_color=BRAND_PURPLE, end_color=BRAND_PURPLE, fill_type="solid")
    ALT_FILL    = PatternFill(start_color="EEF0FF", end_color="EEF0FF", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    TITLE_FONT  = Font(name="Calibri", bold=True, size=14, color=BRAND_PURPLE)
    META_FONT   = Font(name="Calibri", italic=True, size=10, color="888888")

    def _style_header_row(ws, row_num: int, n_cols: int):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _autosize(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 42)

    # ── Sheet 1: Predictions ────────────────────────────────────────────────────
    ws1 = wb.active

    if m["goal"] == "sales_forecast":
        ws1.title = "Sales Forecast"
        ws1["A1"] = "FlowDesk · Sales Forecast"
        ws1["A1"].font = TITLE_FONT
        ws1["A2"] = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  ·  Algorithm: XGBoost Regressor  ·  MAE: ${m.get('mae', 0):,.0f}"
        ws1["A2"].font = META_FONT
        ws1.row_dimensions[3].height = 8

        headers = ["Date", "Actual Revenue ($)", "Predicted Revenue ($)", "Upper Bound ($)", "Lower Bound ($)"]
        for ci, h in enumerate(headers, 1):
            ws1.cell(row=4, column=ci, value=h)
        _style_header_row(ws1, 4, len(headers))
        ws1.freeze_panes = "A5"
        ws1.row_dimensions[4].height = 22

        for ri, pt in enumerate(m.get("forecast_series", []), 5):
            ws1.cell(row=ri, column=1, value=pt.get("date"))
            ws1.cell(row=ri, column=2, value=pt.get("actual"))
            ws1.cell(row=ri, column=3, value=pt.get("predicted"))
            ws1.cell(row=ri, column=4, value=pt.get("upper"))
            ws1.cell(row=ri, column=5, value=pt.get("lower"))
            if ri % 2 == 0:
                for ci in range(1, 6):
                    ws1.cell(row=ri, column=ci).fill = ALT_FILL

    elif m["goal"] == "churn_risk":
        ws1.title = "Churn Risk"
        ws1["A1"] = "FlowDesk · Customer Churn Risk"
        ws1["A1"].font = TITLE_FONT
        ws1["A2"] = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  ·  Algorithm: XGBoost Classifier  ·  AUC: {m.get('auc', 'N/A')}"
        ws1["A2"].font = META_FONT
        ws1.row_dimensions[3].height = 8

        headers = ["Customer ID", "Name", "Churn Risk %", "Revenue at Risk ($)", "Days Inactive", "Risk Level"]
        for ci, h in enumerate(headers, 1):
            ws1.cell(row=4, column=ci, value=h)
        _style_header_row(ws1, 4, len(headers))
        ws1.freeze_panes = "A5"
        ws1.row_dimensions[4].height = 22

        for ri, c in enumerate(m.get("customers", []), 5):
            level = "HIGH" if c["risk"] >= 70 else "MEDIUM" if c["risk"] >= 50 else "LOW"
            ws1.cell(row=ri, column=1, value=c["id"])
            ws1.cell(row=ri, column=2, value=c["name"])
            ws1.cell(row=ri, column=3, value=c["risk"])
            ws1.cell(row=ri, column=4, value=c.get("revenue", 0))
            ws1.cell(row=ri, column=5, value=c.get("days_inactive", 0))
            ws1.cell(row=ri, column=6, value=level)
            if ri % 2 == 0:
                for ci in range(1, 7):
                    ws1.cell(row=ri, column=ci).fill = ALT_FILL

    _autosize(ws1)

    # ── Sheet 2: SHAP Feature Importance ───────────────────────────────────────
    ws2 = wb.create_sheet("SHAP Feature Importance")
    ws2["A1"] = "FlowDesk · SHAP Feature Importance"
    ws2["A1"].font = TITLE_FONT
    ws2["A2"] = "Global mean contribution of each feature to the model's predictions"
    ws2["A2"].font = META_FONT
    ws2.row_dimensions[3].height = 8

    shap_headers = ["Feature (Technical)", "Friendly Name", "Importance %", "Mean SHAP Value"]
    for ci, h in enumerate(shap_headers, 1):
        ws2.cell(row=4, column=ci, value=h)
    _style_header_row(ws2, 4, len(shap_headers))
    ws2.freeze_panes = "A5"

    for ri, feat in enumerate(m.get("shap_global", []), 5):
        ws2.cell(row=ri, column=1, value=feat.get("feature"))
        ws2.cell(row=ri, column=2, value=feat.get("label"))
        ws2.cell(row=ri, column=3, value=feat.get("importance"))
        ws2.cell(row=ri, column=4, value=round(feat.get("raw", 0), 5))
        if ri % 2 == 0:
            for ci in range(1, 5):
                ws2.cell(row=ri, column=ci).fill = ALT_FILL

    _autosize(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    goal_tag = "sales_forecast" if m["goal"] == "sales_forecast" else "churn_risk"
    filename = f"flowdesk_{goal_tag}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Google Sheets Export ───────────────────────────────────────────────────────

@app.post("/api/integrations/sheets/export")
def export_to_sheets(req: SheetsExportRequest):
    """Push model predictions into a Google Spreadsheet via service account."""
    if not _GSPREAD_AVAILABLE:
        raise HTTPException(503, "gspread/google-auth not installed. Run: pip install gspread google-auth")

    cred_row = _get_integration("google_sheets")
    if not cred_row:
        raise HTTPException(400, "Google Sheets not connected. Add your service account JSON in the Integrations hub.")

    try:
        sa_info = json.loads(cred_row["extra"] or "{}")
    except Exception:
        raise HTTPException(400, "Stored service account JSON is invalid. Please reconnect Google Sheets.")

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    try:
        creds = ServiceAccountCreds.from_service_account_info(sa_info, scopes=scopes)
        gc = gspread.authorize(creds)
    except Exception as e:
        raise HTTPException(400, f"Google authentication failed: {e}")

    m = _get_model(req.model_id)

    try:
        if req.spreadsheet_id:
            sh = gc.open_by_key(req.spreadsheet_id)
        else:
            title = f"FlowDesk Predictions — {datetime.utcnow().strftime('%Y-%m-%d')}"
            sh = gc.create(title)
            sa_email = sa_info.get("client_email", "")
            if sa_email:
                sh.share(sa_email, perm_type="user", role="owner", notify=False)
    except Exception as e:
        raise HTTPException(500, f"Could not access/create Google Sheet: {e}")

    try:
        ws = sh.worksheet(req.sheet_name)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=req.sheet_name, rows=600, cols=20)

    rows_written = 0
    if m["goal"] == "sales_forecast":
        ws.append_row(["FlowDesk · Sales Forecast"])
        ws.append_row([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  |  MAE: ${m.get('mae', 0):,.0f}  |  R²: {m.get('r2', 'N/A')}"])
        ws.append_row([])
        ws.append_row(["Date", "Actual Revenue", "Predicted Revenue", "Upper Bound", "Lower Bound"])
        for pt in m.get("forecast_series", []):
            ws.append_row([
                pt.get("date", ""), pt.get("actual", ""),
                pt.get("predicted", ""), pt.get("upper", ""), pt.get("lower", ""),
            ])
            rows_written += 1
        ws.append_row([])
        ws.append_row(["SHAP Feature Importance", "", "", ""])
        ws.append_row(["Feature", "Friendly Name", "Importance %", "Mean SHAP"])
        for feat in m.get("shap_global", []):
            ws.append_row([feat.get("feature", ""), feat.get("label", ""), feat.get("importance", 0), round(feat.get("raw", 0), 5)])

    elif m["goal"] == "churn_risk":
        ws.append_row(["FlowDesk · Customer Churn Risk"])
        ws.append_row([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  |  Accuracy: {m.get('accuracy', 'N/A')}%  |  AUC: {m.get('auc', 'N/A')}"])
        ws.append_row([])
        ws.append_row(["Customer ID", "Name", "Churn Risk %", "Revenue at Risk", "Days Inactive", "Risk Level"])
        for c in m.get("customers", []):
            level = "HIGH" if c["risk"] >= 70 else "MEDIUM" if c["risk"] >= 50 else "LOW"
            ws.append_row([c["id"], c["name"], c["risk"], c.get("revenue", 0), c.get("days_inactive", 0), level])
            rows_written += 1
        ws.append_row([])
        ws.append_row(["SHAP Feature Importance", "", "", ""])
        ws.append_row(["Feature", "Friendly Name", "Importance %", "Mean SHAP"])
        for feat in m.get("shap_global", []):
            ws.append_row([feat.get("feature", ""), feat.get("label", ""), feat.get("importance", 0), round(feat.get("raw", 0), 5)])

    return {
        "url": f"https://docs.google.com/spreadsheets/d/{sh.id}",
        "spreadsheet_id": sh.id,
        "sheet_name": req.sheet_name,
        "rows_written": rows_written,
    }


# ── Slack Webhook ──────────────────────────────────────────────────────────────

@app.post("/api/integrations/slack/send")
def send_slack_message(req: SlackMessageRequest):
    """Send a real Slack message via incoming webhook URL."""
    import requests as _requests
    cred = _get_integration("mcp_slack")
    if not cred or not cred.get("api_key"):
        raise HTTPException(400, "Slack not connected. Add your webhook URL in the Integrations hub.")
    payload: Dict[str, Any] = {"text": req.text}
    if req.blocks:
        payload["blocks"] = req.blocks
    try:
        resp = _requests.post(cred["api_key"], json=payload, timeout=10)
        if not resp.ok:
            raise HTTPException(500, f"Slack webhook returned {resp.status_code}: {resp.text}")
        return {"sent": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Slack send error: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# ── Agent Analytics ─────────────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/agents/analytics")
def get_agent_analytics():
    """
    Returns per-agent 14-day time-series data (run counts + primary KPI metric)
    and a daily total across all agents — used by the agents analytics tab and
    the overview Business Performance Rhythm dropdown.
    """
    today = datetime.utcnow().date()
    cutoff = (today - timedelta(days=13)).isoformat()
    dates = [(today - timedelta(days=i)).isoformat() for i in range(13, -1, -1)]

    con = sqlite3.connect(DB_PATH)

    # Daily run counts per agent from activity log
    run_rows = con.execute(
        "SELECT agent_id, strftime('%Y-%m-%d', created_at) as day, COUNT(*) "
        "FROM agent_activity WHERE created_at >= ? GROUP BY agent_id, day",
        (cutoff,),
    ).fetchall()

    # Primary KPI metric per agent per day
    try:
        metric_rows = con.execute(
            "SELECT agent_id, date, metric_name, AVG(metric_value) "
            "FROM agent_metrics WHERE date >= ? GROUP BY agent_id, date, metric_name",
            (cutoff,),
        ).fetchall()
    except Exception:
        metric_rows = []

    con.close()

    # Index by agent_id → date → run_count
    counts: Dict[str, Dict[str, int]] = {}
    for agent_id, day, cnt in run_rows:
        counts.setdefault(agent_id, {})[day] = cnt

    # Index by agent_id → date → {metric_name: value}
    metrics_idx: Dict[str, Dict[str, Dict[str, float]]] = {}
    for agent_id, date, mname, mval in metric_rows:
        metrics_idx.setdefault(agent_id, {}).setdefault(date, {})[mname] = mval

    # Primary metric label per agent
    primary_metric_label: Dict[str, str] = {
        "marketing":   "Daily Reach",
        "cx":          "Tickets Resolved",
        "sales":       "Leads Scored",
        "seo":         "Keywords Tracked",
        "reputation":  "Reviews Monitored",
        "finance":     "Transactions Reviewed",
        "bi":          "KPIs Tracked",
        "ops":         "Tasks Automated",
        "comms":       "Emails Drafted",
        "compliance":  "Risks Identified",
        "hr":          "Candidates Screened",
        "inventory":   "Items Checked",
    }
    primary_metric_key: Dict[str, str] = {
        "marketing":   "engagement_reach",
        "cx":          "tickets_resolved",
        "sales":       "leads_scored",
        "seo":         "keywords_tracked",
        "reputation":  "reviews_monitored",
        "finance":     "transactions_reviewed",
        "bi":          "kpis_tracked",
        "ops":         "tasks_automated",
        "comms":       "emails_drafted",
        "compliance":  "risks_identified",
        "hr":          "candidates_screened",
        "inventory":   "items_checked",
    }

    # Build per-agent series
    agent_data = {}
    for agent_id, cfg in AGENT_CONFIGS.items():
        pm_key = primary_metric_key.get(agent_id, "tasks_completed")
        series = []
        for d in dates:
            run_count = counts.get(agent_id, {}).get(d, 0)
            metric_val = metrics_idx.get(agent_id, {}).get(d, {}).get(pm_key, 0.0)
            series.append({
                "date": d,
                "label": d[5:],   # MM-DD
                "runs": run_count,
                "tasks": run_count * 3,
                "metric": round(metric_val, 1),
            })
        agent_data[agent_id] = {
            "name": cfg["name"],
            "icon": cfg["icon"],
            "color": cfg["color"],
            "metric_label": primary_metric_label.get(agent_id, "Tasks"),
            "total_runs": sum(counts.get(agent_id, {}).values()),
            "series": series,
        }

    # Daily total across all agents
    daily_totals: Dict[str, int] = {}
    for agent_counts in counts.values():
        for day, cnt in agent_counts.items():
            daily_totals[day] = daily_totals.get(day, 0) + cnt
    daily_total = [{"date": d, "label": d[5:], "runs": daily_totals.get(d, 0)} for d in dates]

    return {"agents": agent_data, "daily_total": daily_total}


@app.get("/api/agents/social/twitter")
def get_twitter_metrics():
    """
    Fetch real Twitter/X tweet engagement metrics using the v2 API Bearer Token.
    Returns per-day engagement (likes + retweets + replies) for the last 20 tweets.
    Requires the twitter_x integration to be connected with a bearer token and username.
    """
    import requests as _requests

    cred = _get_integration("twitter")
    if not cred or not cred.get("api_key"):
        return {"connected": False, "data": []}

    bearer_token = cred["api_key"]
    extra: Dict[str, Any] = {}
    if cred.get("extra"):
        try:
            extra = json.loads(cred["extra"])
        except Exception:
            pass

    username = extra.get("username", "").strip().lstrip("@")
    if not username:
        return {"connected": True, "data": [], "error": "No Twitter username saved. Reconnect and fill in the username field."}

    headers = {"Authorization": f"Bearer {bearer_token}"}

    try:
        # Step 1: resolve username → user ID
        r1 = _requests.get(
            f"https://api.twitter.com/2/users/by/username/{username}",
            headers=headers,
            params={"user.fields": "public_metrics"},
            timeout=10,
        )
        if r1.status_code == 401:
            return {"connected": True, "data": [], "error": "Bearer token invalid or expired."}
        if not r1.ok:
            return {"connected": True, "data": [], "error": f"Twitter API error {r1.status_code}"}

        user = r1.json().get("data", {})
        user_id = user.get("id")
        if not user_id:
            return {"connected": True, "data": [], "error": "Twitter user not found."}

        profile_metrics = user.get("public_metrics", {})

        # Step 2: fetch recent tweets with public_metrics
        r2 = _requests.get(
            f"https://api.twitter.com/2/users/{user_id}/tweets",
            headers=headers,
            params={"tweet.fields": "public_metrics,created_at", "max_results": 20},
            timeout=10,
        )
        if not r2.ok:
            return {"connected": True, "data": [], "error": "Could not fetch tweets."}

        tweets = r2.json().get("data", []) or []

        # Aggregate by calendar date
        daily: Dict[str, dict] = {}
        for tweet in tweets:
            date = (tweet.get("created_at") or "")[:10]
            if not date:
                continue
            m = tweet.get("public_metrics", {})
            if date not in daily:
                daily[date] = {
                    "date": date,
                    "label": date[5:],
                    "likes": 0,
                    "retweets": 0,
                    "replies": 0,
                    "impressions": 0,
                    "posts": 0,
                    "engagement": 0,
                }
            daily[date]["likes"]       += m.get("like_count", 0)
            daily[date]["retweets"]    += m.get("retweet_count", 0)
            daily[date]["replies"]     += m.get("reply_count", 0)
            daily[date]["impressions"] += m.get("impression_count", 0)
            daily[date]["posts"]       += 1
            daily[date]["engagement"]  += (
                m.get("like_count", 0)
                + m.get("retweet_count", 0)
                + m.get("reply_count", 0)
                + m.get("quote_count", 0)
            )

        return {
            "connected": True,
            "username": username,
            "profile_metrics": profile_metrics,
            "data": sorted(daily.values(), key=lambda x: x["date"]),
        }

    except Exception as e:
        return {"connected": True, "data": [], "error": str(e)}
